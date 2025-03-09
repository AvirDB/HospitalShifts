import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import random

# Define departments and workers
DEPARTMENTS = {
    'מיון יולדות': ['עמית', 'גלינה', 'אורים', 'יוני כהן'],
    'מיון נשים': ['גל כהן', 'קוצוק', 'אבוגנים', 'ליבובץ', 'לוין', 'קרטס'],
    'חדר לידה': ['חריש', 'טרסוב', 'מזרחי', 'קציר', 'זמיר', 'אברהמי', 'רביב', 'זומרפרוינד']
}

BEGINNERS = ['עמית', 'גלינה', 'אורים', 'זומרפרוינד']

# Department colors
DEPARTMENT_COLORS = {
    'מיון יולדות': 'FFB6C1',  # Light pink
    'מיון נשים': 'ADD8E6',    # Light blue
    'חדר לידה': '90EE90'      # Light green
}

# Worker's primary department mapping
WORKER_PRIMARY_DEPT = {}
for dept, workers in DEPARTMENTS.items():
    for worker in workers:
        WORKER_PRIMARY_DEPT[worker] = dept

# Worker capabilities
WORKER_CAPABILITIES = {}

# Initialize capabilities for all workers
for dept, workers in DEPARTMENTS.items():
    for worker in workers:
        if worker not in WORKER_CAPABILITIES:
            WORKER_CAPABILITIES[worker] = []
        WORKER_CAPABILITIES[worker].append(dept)
        
        # Workers from חדר לידה can work everywhere
        if dept == 'חדר לידה':
            WORKER_CAPABILITIES[worker].extend(['מיון יולדות', 'מיון נשים'])
        # Workers from מיון נשים can also work in מיון יולדות
        elif dept == 'מיון נשים':
            WORKER_CAPABILITIES[worker].append('מיון יולדות')

# Remove duplicates from capabilities
for worker in WORKER_CAPABILITIES:
    WORKER_CAPABILITIES[worker] = list(set(WORKER_CAPABILITIES[worker]))

class ShiftScheduler:
    def __init__(self):
        self.shifts = {}  # date -> {department -> worker}
        self.worker_shifts_count = {worker: 0 for dept in DEPARTMENTS.values() for worker in dept}
        self.worker_weekend_shifts = {worker: 0 for dept in DEPARTMENTS.values() for worker in dept}
        self.last_shift_date = {worker: None for dept in DEPARTMENTS.values() for worker in dept}
    
    def can_assign_shift(self, worker, date, department):
        # Check if worker can work in this department
        if department not in WORKER_CAPABILITIES[worker]:
            return False
            
        # Check if worker already has a shift on this date
        if date in self.shifts and any(worker == assigned_worker 
                                     for dept, assigned_worker in self.shifts[date].items()):
            return False
            
        # Check if worker had a shift yesterday
        yesterday = date - timedelta(days=1)
        if yesterday in self.shifts and any(worker == assigned_worker 
                                          for dept, assigned_worker in self.shifts[yesterday].items()):
            return False
            
        # Check monthly shift limit (6-7 shifts)
        if self.worker_shifts_count[worker] >= 7:
            return False
            
        # Check weekend shift limit (max 2)
        is_weekend = date.weekday() in [4, 5]  # Friday = 4, Saturday = 5
        if is_weekend and self.worker_weekend_shifts[worker] >= 2:
            return False
            
        # Check beginners constraint
        if worker in BEGINNERS:
            if date in self.shifts:
                for dept, assigned_worker in self.shifts[date].items():
                    if assigned_worker in BEGINNERS:
                        return False
        
        return True
    
    def assign_shift(self, date, department):
        available_workers = [
            worker for worker in WORKER_CAPABILITIES.keys()
            if self.can_assign_shift(worker, date, department)
        ]
        
        if not available_workers:
            return None
            
        # Sort workers by shift count to maintain balance
        available_workers.sort(key=lambda w: (self.worker_shifts_count[w], self.worker_weekend_shifts[w]))
        worker = available_workers[0]
        
        if date not in self.shifts:
            self.shifts[date] = {}
        
        self.shifts[date][department] = worker
        self.worker_shifts_count[worker] += 1
        if date.weekday() in [4, 5]:  # Friday = 4, Saturday = 5
            self.worker_weekend_shifts[worker] += 1
        self.last_shift_date[worker] = date
        
        return worker

def get_cell_color(worker):
    """Get the color for a worker based on their primary department"""
    if worker:
        return DEPARTMENT_COLORS[WORKER_PRIMARY_DEPT[worker]]
    return None

def create_shifts_schedule():
    # Create a date range for December 2024
    start_date = datetime(2024, 12, 1)
    end_date = datetime(2024, 12, 31)
    dates = pd.date_range(start=start_date, end=end_date)
    
    # Initialize scheduler
    scheduler = ShiftScheduler()
    
    # Assign shifts for each date
    for date in dates:
        for department in ['מיון יולדות', 'מיון נשים', 'חדר לידה']:
            scheduler.assign_shift(date, department)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "משמרות דצמבר 2024"
    
    # Set RTL direction for the worksheet
    ws.sheet_view.rightToLeft = True
    
    # Set up headers
    ws['A1'] = 'תאריך'
    ws['B1'] = 'יום'
    ws['C1'] = 'מיון יולדות'
    ws['D1'] = 'מיון נשים'
    ws['E1'] = 'חדר לידה'
    
    # Hebrew days mapping
    hebrew_days = {
        0: 'שני',
        1: 'שלישי',
        2: 'רביעי',
        3: 'חמישי',
        4: 'שישי',
        5: 'שבת',
        6: 'ראשון'
    }
    
    # Fill dates and shifts
    for idx, date in enumerate(dates, start=2):
        ws[f'A{idx}'] = date.strftime('%d/%m/%Y')
        ws[f'B{idx}'] = hebrew_days[date.weekday()]
        
        # Fill shifts and color cells based on worker's department
        if date in scheduler.shifts:
            for col, dept in zip(['C', 'D', 'E'], ['מיון יולדות', 'מיון נשים', 'חדר לידה']):
                worker = scheduler.shifts[date].get(dept, '')
                cell = ws[f'{col}{idx}']
                cell.value = worker
                if worker:
                    cell.fill = PatternFill(start_color=get_cell_color(worker),
                                          end_color=get_cell_color(worker),
                                          fill_type='solid')
        
        # Color weekends (Friday and Saturday)
        if date.weekday() in [4, 5]:  # Friday = 4, Saturday = 5
            for col in ['A', 'B']:  # Only color date and day columns
                ws[f'{col}{idx}'].fill = PatternFill(start_color='FFE6E6', 
                                                    end_color='FFE6E6',
                                                    fill_type='solid')
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Center align all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Make headers bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title="סיכום משמרות")
    
    # Set RTL direction for the summary worksheet
    summary_ws.sheet_view.rightToLeft = True
    
    summary_ws['A1'] = 'שם העובד'
    summary_ws['B1'] = 'מחלקה'
    summary_ws['C1'] = 'סה"כ משמרות'
    summary_ws['D1'] = 'משמרות סופ"ש'
    
    # Fill summary data
    row = 2
    for dept, workers in DEPARTMENTS.items():
        for worker in workers:
            summary_ws[f'A{row}'] = worker
            summary_ws[f'B{row}'] = dept
            summary_ws[f'C{row}'] = scheduler.worker_shifts_count[worker]
            summary_ws[f'D{row}'] = scheduler.worker_weekend_shifts[worker]
            
            # Color the row based on department
            for col in ['A', 'B', 'C', 'D']:
                cell = summary_ws[f'{col}{row}']
                cell.fill = PatternFill(start_color=DEPARTMENT_COLORS[dept],
                                      end_color=DEPARTMENT_COLORS[dept],
                                      fill_type='solid')
            row += 1
    
    # Set column widths for summary sheet
    summary_ws.column_dimensions['A'].width = 15
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 15
    summary_ws.column_dimensions['D'].width = 15
    
    # Center align all cells in summary sheet
    for row in summary_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Make headers bold in summary sheet
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    
    # Save the workbook
    wb.save('hospital_shifts_december_2024.xlsx')

if __name__ == "__main__":
    create_shifts_schedule() 